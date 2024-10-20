import os
import openpyxl
from pathlib import Path

_FILENAME = 'etfs.xlsx'


class ETFLoader:
    def __init__(self):
        self.etf_issuers = []
        self.etf_codes = []
        self.etf_names = []
        self.load_xlsx()

    def load_xlsx(self):
        project_path = os.getcwd()
        file_path = Path(project_path) / 'etf_issuer' / _FILENAME
        workbook = openpyxl.load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            self.etf_issuers.append(sheet_name)
            sheet = workbook[sheet_name]
            codes, names = zip(*[(row[0], row[1]) for row in sheet.iter_rows(values_only=True)])
            self.etf_codes.append(codes)
            self.etf_names.append(names)

    def get_titles(self, index):
        return [f'{code} {name}' for code, name in zip(self.etf_codes[index], self.etf_names[index])]

    def has_etf(self, etf):
        return any(etf in etfs for etfs in self.etf_codes)
